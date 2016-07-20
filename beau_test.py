#!/usr/bin/env python
# encoding: utf-8

"""
@version: ??
@author: kabi
@license: Apache Licence 
@file: beau_test.py
@time: 2016/7/1 0001 上午 12:13
"""

from bs4 import BeautifulSoup
import urllib
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import titlecase
kabi = []
def eat_spider():
	page_num = 0


	for i in range(1):
		html_doc = 'https://www.douban.com/subject/all?cat_id=1000&start=' + str(20*i)
		page = urllib.urlopen(html_doc)
		#html = page.read()
		soup = BeautifulSoup(page,'html.parser')
		test = soup.findAll('li',{'class':"thing-item"})
		test_li = soup.findAll('span',{'class':"favs"})
		for i in range(0,len(test)):
			test_title = test[i].a['title']
			test_web = test[i].a['href']
			test_pic = test[i].img['src']
			test_like = test_li[i].get_text()
			xixi = test_title + ' web:' + test_web +	' pic:' + test_pic + ' like:' +test_like
			kabi.append([test_title,test_web,test_pic,test_like])
			print(kabi)

		i += 1
	write_excel()

def write_excel():
	wb = load_workbook(filename='haha.xlsx')
	ws = wb['hsdf']
	#测试用 后面循环~~~
	for i in range(0,len(kabi)):
		ws.cell(row = i+1,column = 1).value = kabi[i][0]
		ws.cell(row = i+1,column = 2).value = kabi[i][1]
		ws.cell(row = i+1,column = 3).value = kabi[i][2]
		ws.cell(row = i+1,column = 4).value = kabi[i][3]

	wb.save(filename='haha.xlsx')

def main():
	eat_spider()
	#write_excel()

if __name__ == '__main__':
	main()